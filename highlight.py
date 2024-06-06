import openpyxl
import glob
import re
from openpyxl.styles import PatternFill
from openpyxl.cell.text import InlineFont
from openpyxl.cell.rich_text import TextBlock, CellRichText

INPUT_DIR = "input"
OUTPUT_DIR = "output"
dob_keyword = ['dob', 'date of birth']
yellow = 'FFFF00'


def find_dob_pattern(text):
    regex = re.compile(r"(?:^|\b)(" + "|".join(dob_keyword) + r")(?:\b|$)", re.IGNORECASE)
    keyword_found = [m.span() for m in regex.finditer(text)]
    return keyword_found


def find_phone_pattern(text):
    regex = re.compile(r"(?:^|\b|\s)(\d{3}[-\.\s]??\d{3}[-\.\s]??\d{4}|\(\d{3}\)\s*\d{3}[-\.\s]??\d{4})(?:\b|$)")

    keyword_found = [m.span() for m in regex.finditer(text)]
    return keyword_found


def highlight(text, match_span_list):
    start_pos = 0
    rich_text_cell = CellRichText()
    red = InlineFont(color='00FF0000')

    for span in match_span_list:
        span_start, span_end = span
        if start_pos < span_start:
            rich_text_cell.append(text[start_pos: span_start])
        rich_text_cell.append(TextBlock(red, text[span_start: span_end]))

        start_pos = span_end

    rich_text_cell.append(text[start_pos:])
    return rich_text_cell


def excel_process(file_name):
    book = openpyxl.load_workbook(INPUT_DIR + "\\" + file_name)
    sheet = book.active
    row_number = 1
    match = 0

    for row in sheet.iter_rows():
        if row_number == 1:
            row_number += 1
            continue
        cell_value = sheet.cell(row=row_number, column=10).value
        dob_match = find_dob_pattern(str(cell_value))
        phone_match = find_phone_pattern(str(cell_value))
        all_match = dob_match + phone_match

        if all_match:
            for cell in row:
                cell.fill = PatternFill(start_color="00FFFF00", fill_type = "solid")
            match += 1
            highlighted_text = highlight(str(cell_value), all_match)

            sheet.cell(row=row_number, column=10).value = highlighted_text


        row_number += 1
    book.save(OUTPUT_DIR + "\\" + file_name)


def get_files_from_input_dir():
    dir_list = glob.glob(INPUT_DIR + "\\" + '*.xlsx')
    for file in dir_list:
        file_name = file.split("\\")[1]
        excel_process(file_name)


get_files_from_input_dir()
